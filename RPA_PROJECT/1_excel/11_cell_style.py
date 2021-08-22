from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("sample2.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]


# a열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 1번째 줄의 높이를 50으로 설정
ws.row_dimensions[1].height = 50

## 글자 스타일
# 빨간색, 이탤릭
a1.font = Font(color = "FF0000", italic = True, bold = True)    # 빨간색, 이태릭(기울임), 굵게
b1.font = Font(color = "CC33FF", name = "Arial", strike = True) # 폰트를 Arial, 취소선(strike)
c1.font = Font(color = "0000FF", size = 20, underline="single") # 글자 크기를 20, 밑줄 적용

## 셀의 테두리
thin_border = Border(left = Side(style = "thin"), right = Side(style = "thin"), top = Side(style = "thin"), bottom = Side(style = "thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border


for row in ws.rows :
    for cell in row :
        # 각 셀을 센터로 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if cell.column == 1 :
            continue
        
        if isinstance(cell.value, int) and cell.value > 90 :
            # 셀의 배경색을 지정
            cell.fill = PatternFill(fgColor="00FF00", fill_type = "solid")
            # 셀의 폰트를 지정
            cell.font = Font(color = "FF0000")

## 틀 고정
ws.freeze_panes = "A2"

wb.save("sample2_text.xlsx")
