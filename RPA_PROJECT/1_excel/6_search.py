from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string  ## 좌표 정보를 가져오기 위한 모듈

from random import *
## 새로운 워크북 작성
# wb = Workbook()

## 기존의 워크북을 불러옴
wb = load_workbook("sample2.xlsx")
ws = wb.active

## row 를 순회하며 값을 찾기
for row in ws.iter_rows(max_row = 5): # 첫번째 열의 값만 찾음ㄴ
    for cell in row :
        print(cell.value)
        if cell.value == "영어" :
            print("찾음 :", cell.value)