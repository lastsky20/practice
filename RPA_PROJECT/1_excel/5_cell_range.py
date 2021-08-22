from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string  ## 좌표 정보를 가져오기 위한 모듈

from random import *    
## 새로운 워크북을 생성
wb = Workbook()
ws = wb.active


## 한 줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11) :
    ws.append([i, randint(0, 100), randint(0, 100)])

## 한개의 컬럼 데이터 가져오기
col_B = ws["B"] # B 컬럼의 값만 가져오기

# for cell in col_B :
#     print(cell.value, end = " ")

## 여러 컬럼의 데이터 가져오기
col_range = ws["B:C"]   # B, C 컬럼의 데이터 가져오기

# print(col_range)

# for cols in col_range :   # 컬럼의 데이터(튜플) 만 가져옴
#     # print(cols)
#     for cell in cols :    # 컬럼의 데이터 출력
#         print(cell.value, end = " ")
## 

## 첫 번째 row 만 가져오기
# title = ws[1]
# for name in title :
#     print(name.value, end = " ")


## 여러 row 데이터 가져오기
title = ws[1:ws.max_row]    # 1번째 row 부터 끝까지
for names in title :
    for name in names :
        print(name.value, end = " " )
    print()

## 여러 row 데이터의 좌료값 가져오기
# title = ws[1:ws.max_row]    # 1번째 row 부터 끝까지
# for names in title :
#     for name in names :
#         xy = coordinate_from_string(name.coordinate)
#         print(xy)   # 한줄로 각각의 좌표 값 출력
#         # print(xy[0], end = "") # 한줄로 하나로된 좌표값 출력
#         # print(xy[1], end = " ")   
#     print()


## 전체 row
# print(tuple(ws.rows))

## 전체 column
# print(tuple(ws.columns))

# # 전체 row 의 데이터 가져오기
# for row in tuple(ws.rows) :
#     print(row[0].value)

# for column in tuple(ws.columns) :
#     print(column[1].value)


## 전체 row 데이터 가져오기
# for row in ws.iter_rows() :
#     print(row[1].value)
# for column in ws.iter_cols():
#     print(column[0].value)

## 셀 범위를 지정하여 데이터 가져오기
# for row in ws.iter_rows(min_row = 1, max_row = 11, min_col = 1, max_col = 3) :  # 좌에서 우로 데이터를 가져옴
#     print(row[0].value, row[1].value)

for col in ws.iter_cols(min_row = 1, max_row = 5, min_col = 1, max_col = 3) :   # 위에서 아래로 데이터를 가져옴
    print(col[0].value, col[1].value)


wb.save("sample2.xlsx")

