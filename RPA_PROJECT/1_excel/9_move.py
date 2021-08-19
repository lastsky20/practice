from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string  ## 좌표 정보를 가져오기 위한 모듈

from random import *
## 새로운 워크북 작성
# wb = Workbook()

## 기존의 워크북을 불러옴
wb = load_workbook("sample2.xlsx")
ws = wb.active

# 번호 영어 수학
# 번호 영어 국어 수학

# ws.move_range("B1:C11", rows = 0, cols = 1) # 이동하려는 범위 "B1:C11" 기준으로 줄은 그대로, 1개 열을 이동
# ws.move_range("B1:C11", rows = 3, cols = 0) # 이동하려는 범위 "B1:C11" 기준으로 3줄은 아래로 이동
ws.move_range("C1:C11", rows = 5, cols = -1)



## 터미널에 엑셀의 내용을 임시출력
print("max_column = ", ws.max_column)
print("max_row = ", ws.max_row)
for x in range(1, ws.max_row+1) :
    for y in range(1, ws.max_column+1) :
        print(ws.cell(column = y, row = x).value, end = " " )
    print()


