from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook("sample_test2.xlsx")
ws = wb.active

# ws.unmerge_cells("B2:D2")
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

wb.save("Sample_test2.xlsx")

