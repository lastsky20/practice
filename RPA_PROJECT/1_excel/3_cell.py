from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active  # 활성화된 시트 사용
ws.title = "TestSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6


# 직접 접근
print(ws["A1"]) # 셀의 정보를 출력
print(ws["A1"].value)   # 셀의 값을 출력
print(ws["A10"].value)  # 값이 없을땐 none 출력

# 셀의 column, row 값으로 접근
print(ws.cell(column = 1, row = 1).value)
print(ws.cell(column = 2, row = 1).value)

# value 값을 지정하여 데이터 입력
w = ws.cell(column = 3, row = 1, value = 10)    # column, row 값을 지정하여 값 출력 == ws["C1"] = 10
print(ws.cell(column = 3, row = 1).value)   
print(w.value)  # 변수를 지정하여 값 출력 == w["C1"].value


i = 1
for x in range(1, 11) : 
    for y in range(1, 11) :
        # ws.cell(column = x, row = y, value = randint(1, 100))
        ws.cell(row = x, column = y, value = i)
        i += 1


wb.save("TestBook.xlsx")