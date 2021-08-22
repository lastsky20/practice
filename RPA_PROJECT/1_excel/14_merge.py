from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active


## 셀 병합하기
ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"
# 셀 병합후 센터 정렬
ws["B2"].alignment = Alignment(horizontal="center", vertical="center")

wb.save("Sample_test2.xlsx")

