from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string  ## 좌표 정보를 가져오기 위한 모듈

from random import *
## 새로운 워크북 작성
# wb = Workbook()

## 기존의 워크북을 불러옴
wb = load_workbook("sample2.xlsx")
ws = wb.active




# ws.delete_rows(8) # 8번째 "줄" 삭제
# ws.delete_rows(8, 3) # 8번째 "줄" 부터 "3줄" 삭제
# ws.delete_cols(2) # B의 열 삭제
ws.delete_cols(2, 2)    # B열부터 2열을 삭제

print("max_column = ", ws.max_column)
print("max_row = ", ws.max_row)
for x in range(1, ws.max_row+1) :
    for y in range(1, ws.max_column+1) :
        print(ws.cell(column = y, row = x).value, end = " " )
    print()





