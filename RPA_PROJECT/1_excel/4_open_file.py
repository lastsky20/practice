from openpyxl import Workbook
from openpyxl import load_workbook


wb = load_workbook("sample.xlsx")
ws = wb.active


## 셀의 데이터를 알고 있을 때, 
i = 1
for x in range(1, 11) :
    for y in range(1, 11) : 
        ws.cell(row = x, column = y, value = i)
        # print(ws.cell(row = x, column = y).value, end = " ")
        i += 1
    print()

## 셀의 데이터를 모를 때 / ws.max_row / ws.max_column
for x in range(1, ws.max_row + 1) :
    for y in range(1, ws.max_column + 1) :
        print(ws.cell(row = x, column = y).value, end = " ")
    print()


