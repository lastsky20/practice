from typing import Counter
from openpyxl import Workbook
from openpyxl.descriptors.base import Integer
from openpyxl.styles import Alignment
from openpyxl import load_workbook


# wb = Workbook()
wb = load_workbook("scores.xlsx")
ws = wb.active


# ## 엑셀 데이터 입력부분 
# title = ["학번", "출석", "퀴즈1", "퀴즈2", 
#         "중간고사", "기말고사", "프로젝트", "줄바꿈",
#         1,10,8,5,14,26,12,"줄바꿈",
#         2,7,3,7,15,24,18,"줄바꿈",
#         3,9,5,8,8,12,4,"줄바꿈",
#         4,7,8,7,17,21,18,"줄바꿈",
#         5,7,8,7,16,25,15,"줄바꿈",
#         6,3,5,8,8,17,0,"줄바꿈",
#         7,4,9,10,16,27,18,"줄바꿈",
#         8,6,6,6,15,19,17,"줄바꿈",
#         9,10,10,9,19,30,19,"줄바꿈",
#         10,9,8,8,20,25,20]

# ## 엑셀 데이터 입력 풀이
ws.append = (("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))
scores = [
    (1,10,8,5,14,26,12),
    (2,7,3,7,15,24,18),
    (3,9,5,8,8,12,4),
    (4,7,8,7,17,21,18),
    (5,7,8,7,16,25,15),
    (6,3,5,8,8,17,0),
    (7,4,9,10,16,27,18),
    (8,6,6,6,15,19,17),
    (9,10,10,9,19,30,19),
    (10,9,8,8,20,25,20)
]

# i = 1
# j = 1
# for name in title :
#     # print(i)
#     # ws.append(j, )
#     if name == "줄바꿈" :
#         print("w")
#         j = j + 1
#         i = 1
#         continue
#     ws.cell(column = i, row = j, value = name).alignment = Alignment(horizontal="center", vertical="center")
#     i = i + 1

## 퀴즈 풀이 부분

# 퀴즈2 칼럼의 데이터를 모두 10으로 수정
for i in range(2, 12) :
    ws.cell(column=4, row=i, value = 10)

# 총점 계산 부분
ws["H1"] ="총점"
total = 0
for i in range(2, 12) :
    for j in range(2, 8) :
        total = total + ws.cell(column=j, row=i).value
    ws.cell(column=j+1, row=i, value=int(total))
    total = 0


# 성적 계산 부분
ws["I1"] ="성적"
for i in range(2, 12) :
        if ws.cell(column=2, row=i).value < 5 :
            ws.cell(column=9, row=i, value = "F")
        elif ws.cell(column=8, row=i).value > 90 :
            ws.cell(column=9, row=i, value="A")
        elif ws.cell(column=8, row=i).value > 80 :
            ws.cell(column=9, row=i, value="B")
        elif ws.cell(column=8, row=i).value > 70 :
            ws.cell(column=9, row=i, value="C")
        else :
            ws.cell(column=9, row=i, value="D")
wb.save("scores.xlsx")