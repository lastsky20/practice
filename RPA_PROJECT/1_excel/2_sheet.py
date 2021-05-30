from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()  # 새로운 시트 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "1361aa"  # html rgb 컬러값 / https://www.w3schools.com/colors

ws1 = wb.create_sheet("MySheet2")   # 주어진 이름으로 시트 생성
ws2 = wb.create_sheet("MySheet3", 2)    # 2번째 시트에 생성, 

new_ws = wb["MySheet3"] # Dictionary 형태로 시트에 접근 가능
# new_ws.  으로 해당 시트 내용에 접근 가능

print(wb.sheetnames)    # 모든 시트 확인

new_ws["A1"] = "test"   # 첫번째 셀에 데이터 입력

target = wb.copy_worksheet(new_ws)  # 시트 복사
target.title = "Copied Sheet3"

ws3 = wb.copy_worksheet(ws1)
ws3.title = "Copied Sheet1"

wb.save("sample.xlsx")