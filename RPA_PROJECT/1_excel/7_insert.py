from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string  ## 좌표 정보를 가져오기 위한 모듈

from random import *
## 새로운 워크북 작성
# wb = Workbook()

## 기존의 워크북을 불러옴
wb = load_workbook("sample2.xlsx")
ws = wb.active

# ws.insert_rows(8, 5) # 8번째 줄부터 5번의 삽입
# ws.insert_cols(2, 5) # B번째 열부터 5번의 삽입
# ws.insert_cols(2) # B번째 열이 삽입됨



wb.save("sample2.xlsx")