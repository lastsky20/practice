from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image



wb = Workbook()
ws = wb.active

## 이미지 삽입
# img 객체 생성
img = Image("capture1.jpg")

# 이미지의 사이즈 지정
img.height = 50
img.width = 50

# D3 위치에 이미지 삽입
ws.add_image(img, "D3")

wb.save("sample_test3.xlsx")