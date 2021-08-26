from typing import Counter
from openpyxl import Workbook
from openpyxl import load_workbook


wb = Workbook()
ws = wb.active



# ## 엑셀 데이터 입력부분 
title = (("학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"))
scores = [(1,10,8,5,14,26,12),
        (2,7,3,7,15,24,18),
        (3,9,5,8,8,12,4),
        (4,7,8,7,17,21,18),
        (5,7,8,7,16,25,15),
        (6,3,5,8,8,17,0),
        (7,4,9,10,16,27,18),
        (8,6,6,6,15,19,17),
        (9,10,10,9,19,30,19),
        (10,9,8,8,20,25,20)]

ws.append(title)

for score in scores :
    ws.append(score)

# 퀴즈2 칼럼의 데이터를 모두 10으로 수정
for idx, cell in enumerate(ws["D"]) :
    if idx == 0 :
        continue
    cell.value = 10

ws["H1"] = "총점"
ws["I1"] = "학점"

## 총점 계산
for idx, score in enumerate(scores, start = 2) :
    sum_val = sum(score[1:]) - score[3] + 10
    # ws.cell(row=idx, column=8).value = "=SUM(B{}:G{})".format(idx, idx)
    ws.cell(row=idx, column=8).value = sum_val
    # B1:G1
    # B2:G2 ...
    
    ## 학점 계산
    grade = None

    if ws.cell(row=idx, column=2).value < 5 :
        grade = "F"
    else :
        if sum_val >= 90 :
            grade = "A"
        elif sum_val >= 80 :
            grade = "B"
        elif sum_val >= 70 :
            grade = "C"
        else :
            grade = "D"
    ws.cell(row=idx, column=9).value = grade
    


## 총점 계산


wb.save("quiz_result.xlsx")


